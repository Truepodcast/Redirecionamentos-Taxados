from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)  # Permite chamadas CORS do navegador

EXCEL_PATH = "cliques.xlsx"

def salvar_no_excel(origem, link):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cliques"
        ws.append(["Data/Hora", "Origem", "Link"])
    else:
        wb = load_workbook(EXCEL_PATH)
        if "Cliques" in wb.sheetnames:
            ws = wb["Cliques"]
        else:
            ws = wb.active

    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), origem, link])
    wb.save(EXCEL_PATH)

@app.route('/')
def index():
    return send_from_directory('.', 'landing_lista_com_rastreamento.html')

@app.route('/registrar', methods=['OPTIONS', 'POST', 'GET'])
def registrar():
    if request.method == 'OPTIONS':
        return jsonify({'message': 'OK'}), 200

    # Permite rastreamento via GET também
    if request.method == 'GET':
        origem = request.args.get("origem")
        link = request.args.get("link", "https://www.youtube.com/live/ri110bt97Vs?si=EhianlKzj5KXkXQa")
        if not origem:
            return jsonify({'erro': 'Origem não definida'}), 400
        salvar_no_excel(origem, link)
        return send_from_directory('.', 'redirect.html')

    data = request.get_json()
    origem = data.get("origem")
    link = data.get("link")

    if not origem or not link:
        return jsonify({'erro': 'Dados inválidos'}), 400

    salvar_no_excel(origem, link)
    print(f"[{datetime.now().isoformat()}] Clique registrado - Origem: {origem}, Link: {link}")
    return jsonify({'message': 'Registrado com sucesso!'}), 201

if __name__ == '__main__':
    app.run(debug=True)